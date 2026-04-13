import numpy as np
import pandas as pd
from collections import defaultdict
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font

# Load the ID-Score data
df = pd.read_excel('/mnt/user-data/outputs/id_scores.xlsx')
id_to_score = dict(zip(df['ID'], df['Score']))
all_ids = list(df['ID'].values)

# Group IDs by score for easier selection
ids_by_score = defaultdict(list)
for id_val, score in id_to_score.items():
    ids_by_score[score].append(id_val)

# Parameters
n_subsamples = 1000
subsample_size = 20
expected_frequency = (n_subsamples * subsample_size) / len(all_ids)  # 200

print(f"Target: {n_subsamples} subsamples of {subsample_size} IDs each")
print(f"Expected frequency per ID: {expected_frequency}")

# Generate target average scores uniformly distributed
# Use more granular targets to achieve true uniformity
np.random.seed(42)
target_scores = np.linspace(0.1, 0.9, n_subsamples)  # From 0.1 to 0.9
np.random.shuffle(target_scores)

# Track usage count for each ID
usage_count = defaultdict(int)

# Store all subsamples
subsamples = []
subsample_averages = []

def select_subsample_greedy(target_avg, usage_count):
    """
    Greedily construct a subsample that hits the target average
    while balancing ID usage.
    """
    selected = []
    selected_scores_sum = 0
    
    for position in range(subsample_size):
        remaining = subsample_size - position
        needed_sum = target_avg * subsample_size - selected_scores_sum
        ideal_next = needed_sum / remaining
        
        # Find available IDs (not in current subsample)
        available_by_score = {}
        for score, id_list in ids_by_score.items():
            available_ids = [id_val for id_val in id_list if id_val not in selected]
            if available_ids:
                available_by_score[score] = available_ids
        
        if not available_by_score:
            break
            
        # Find the score closest to ideal
        available_scores = list(available_by_score.keys())
        score_diffs = [abs(s - ideal_next) for s in available_scores]
        
        # Get top 2-3 closest scores for diversity
        sorted_indices = np.argsort(score_diffs)
        candidate_scores = [available_scores[i] for i in sorted_indices[:min(3, len(sorted_indices))]]
        
        # Among these scores, select ID with lowest usage
        best_id = None
        best_usage = float('inf')
        best_score = None
        
        for score in candidate_scores:
            for id_val in available_by_score[score]:
                usage_penalty = usage_count[id_val]
                # Add small random factor to break ties
                usage_with_noise = usage_penalty + np.random.uniform(-0.5, 0.5)
                if usage_with_noise < best_usage:
                    best_usage = usage_with_noise
                    best_id = id_val
                    best_score = score
        
        if best_id is not None:
            selected.append(best_id)
            selected_scores_sum += best_score
    
    return selected

# Generate subsamples with retry logic for extreme targets
for idx, target_avg in enumerate(target_scores):
    if (idx + 1) % 100 == 0:
        print(f"Generating subsample {idx + 1}/{n_subsamples}...")
    
    # Try to generate a subsample multiple times, keep the best
    best_subsample = None
    best_error = float('inf')
    
    for attempt in range(5):  # Try up to 5 times
        subsample = select_subsample_greedy(target_avg, usage_count)
        
        if len(subsample) == subsample_size:
            actual_avg = np.mean([id_to_score[id_val] for id_val in subsample])
            error = abs(actual_avg - target_avg)
            
            if error < best_error:
                best_error = error
                best_subsample = subsample
            
            # If error is small enough, accept immediately
            if error < 0.02:
                break
    
    # Use the best subsample found
    subsamples.append(best_subsample)
    actual_avg = np.mean([id_to_score[id_val] for id_val in best_subsample])
    subsample_averages.append(actual_avg)
    
    # Update usage counts
    for id_val in best_subsample:
        usage_count[id_val] += 1

print("\nGeneration complete! Verifying conditions...\n")

# ============ VERIFICATION ============

print("CONDITION 1: Uniform distribution of average scores")
print(f"Min average: {min(subsample_averages):.4f}")
print(f"Max average: {max(subsample_averages):.4f}")
print(f"Mean average: {np.mean(subsample_averages):.4f}")
print(f"Std dev of averages: {np.std(subsample_averages):.4f}")

# Check uniformity using histogram bins
hist, bins = np.histogram(subsample_averages, bins=10)
print(f"\nDistribution across 10 bins:")
for i, count in enumerate(hist):
    print(f"  Bin {i+1} [{bins[i]:.2f}-{bins[i+1]:.2f}]: {count} subsamples")

# Calculate uniformity metric (Chi-square test)
expected_per_bin = n_subsamples / 10
chi_square = sum((count - expected_per_bin)**2 / expected_per_bin for count in hist)
print(f"\nChi-square statistic (lower is more uniform): {chi_square:.2f}")
print(f"Perfect uniformity would be: 0.00")

# Condition 2: Check frequency distribution of IDs
all_ids_used = [id_val for subsample in subsamples for id_val in subsample]
from collections import Counter
id_frequencies = Counter(all_ids_used)

frequencies = list(id_frequencies.values())
print(f"\nCONDITION 2: Uniform distribution of ID frequencies")
print(f"Expected frequency per ID: {expected_frequency:.1f}")
print(f"Min frequency: {min(frequencies)}")
print(f"Max frequency: {max(frequencies)}")
print(f"Mean frequency: {np.mean(frequencies):.2f}")
print(f"Std dev of frequencies: {np.std(frequencies):.2f}")
print(f"Range: {max(frequencies) - min(frequencies)}")
print(f"Coefficient of variation: {np.std(frequencies) / np.mean(frequencies):.3f}")

# ============ SAVE RESULTS ============

wb = Workbook()

# Sheet 1: Subsamples
ws1 = wb.active
ws1.title = "Subsamples"
ws1['A1'] = 'Subsample_ID'
ws1['B1'] = 'Target_Avg'
ws1['C1'] = 'Actual_Avg'
ws1['D1'] = 'Error'
for i in range(subsample_size):
    ws1.cell(1, i+5, f'ID_{i+1}')

for col in range(1, subsample_size + 5):
    ws1.cell(1, col).font = Font(bold=True)

for idx, (subsample, target, actual) in enumerate(zip(subsamples, target_scores, subsample_averages), 1):
    ws1.cell(idx+1, 1, idx)
    ws1.cell(idx+1, 2, round(target, 4))
    ws1.cell(idx+1, 3, round(actual, 4))
    ws1.cell(idx+1, 4, round(abs(actual - target), 4))
    for i, id_val in enumerate(subsample):
        ws1.cell(idx+1, i+5, id_val)

# Sheet 2: ID Frequency Summary
ws2 = wb.create_sheet("ID_Frequencies")
ws2['A1'] = 'ID'
ws2['B1'] = 'Score'
ws2['C1'] = 'Frequency'
ws2['D1'] = 'Deviation_from_Expected'

for col in range(1, 5):
    ws2.cell(1, col).font = Font(bold=True)

for idx, id_val in enumerate(sorted(all_ids), 1):
    ws2.cell(idx+1, 1, id_val)
    ws2.cell(idx+1, 2, id_to_score[id_val])
    ws2.cell(idx+1, 3, id_frequencies[id_val])
    ws2.cell(idx+1, 4, id_frequencies[id_val] - expected_frequency)

# Sheet 3: Summary Statistics
ws3 = wb.create_sheet("Summary")
ws3['A1'] = 'Metric'
ws3['B1'] = 'Value'
ws3['A1'].font = Font(bold=True)
ws3['B1'].font = Font(bold=True)

summary_data = [
    ['Total Subsamples', n_subsamples],
    ['Subsample Size', subsample_size],
    ['Total IDs', len(all_ids)],
    ['Expected Frequency per ID', expected_frequency],
    ['', ''],
    ['CONDITION 1: Subsample Averages', ''],
    ['Min Average', round(min(subsample_averages), 4)],
    ['Max Average', round(max(subsample_averages), 4)],
    ['Mean Average', round(np.mean(subsample_averages), 4)],
    ['Std Dev', round(np.std(subsample_averages), 4)],
    ['Chi-square (uniformity)', round(chi_square, 2)],
    ['', ''],
    ['CONDITION 2: ID Frequencies', ''],
    ['Min Frequency', min(frequencies)],
    ['Max Frequency', max(frequencies)],
    ['Mean Frequency', round(np.mean(frequencies), 2)],
    ['Std Dev', round(np.std(frequencies), 2)],
    ['Range', max(frequencies) - min(frequencies)],
    ['Coefficient of Variation', round(np.std(frequencies) / np.mean(frequencies), 3)],
]

for idx, (metric, value) in enumerate(summary_data, 1):
    ws3.cell(idx, 1, metric)
    ws3.cell(idx, 2, value)

wb.save('/mnt/user-data/outputs/subsamples_results_v2.xlsx')

# ============ VISUALIZATIONS ============

fig, axes = plt.subplots(2, 2, figsize=(14, 10))

# Plot 1: Distribution of subsample averages
axes[0, 0].hist(subsample_averages, bins=30, edgecolor='black', alpha=0.7, color='steelblue')
axes[0, 0].axhline(n_subsamples/30, color='red', linestyle='--', 
                   label=f'Uniform expectation ({n_subsamples/30:.1f})')
axes[0, 0].set_xlabel('Average Score')
axes[0, 0].set_ylabel('Number of Subsamples')
axes[0, 0].set_title('Condition 1: Distribution of Subsample Averages')
axes[0, 0].legend()
axes[0, 0].grid(True, alpha=0.3)

# Plot 2: Target vs Actual averages
axes[0, 1].scatter(target_scores, subsample_averages, alpha=0.3, s=10)
axes[0, 1].plot([0, 1], [0, 1], 'r--', label='Perfect match')
axes[0, 1].set_xlabel('Target Average')
axes[0, 1].set_ylabel('Actual Average')
axes[0, 1].set_title('Target vs Actual Average Scores')
axes[0, 1].legend()
axes[0, 1].grid(True, alpha=0.3)

# Plot 3: ID frequency distribution
axes[1, 0].hist(frequencies, bins=30, edgecolor='black', alpha=0.7, color='green')
axes[1, 0].axvline(expected_frequency, color='red', linestyle='--', 
                   label=f'Expected ({expected_frequency:.0f})')
axes[1, 0].set_xlabel('Frequency (times ID appears)')
axes[1, 0].set_ylabel('Number of IDs')
axes[1, 0].set_title('Condition 2: Distribution of ID Frequencies')
axes[1, 0].legend()
axes[1, 0].grid(True, alpha=0.3)

# Plot 4: Frequency by ID score
freq_by_score = defaultdict(list)
for id_val, freq in id_frequencies.items():
    score = id_to_score[id_val]
    freq_by_score[score].append(freq)

scores = sorted(freq_by_score.keys())
avg_freqs = [np.mean(freq_by_score[s]) for s in scores]
axes[1, 1].bar(scores, avg_freqs, width=0.15, alpha=0.7, color='orange')
axes[1, 1].axhline(expected_frequency, color='red', linestyle='--', 
                   label=f'Expected ({expected_frequency:.0f})')
axes[1, 1].set_xlabel('ID Score')
axes[1, 1].set_ylabel('Average Frequency')
axes[1, 1].set_title('Average ID Frequency by Score Value')
axes[1, 1].legend()
axes[1, 1].grid(True, alpha=0.3)

plt.tight_layout()
plt.savefig('/mnt/user-data/outputs/verification_plots_v2.png', dpi=150, bbox_inches='tight')

print(f"\n✓ Results saved to 'subsamples_results_v2.xlsx'")
print(f"✓ Visualizations saved to 'verification_plots_v2.png'")
